import json
import os

import openpyxl as op
import pandas as pd
import requests

TELEGRAM_SHEET = "TELEGRAM"
VK_SHEET = "ВКОНТАКТЕ"


def load_sheet(sheet_id: str) -> None:
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    response = requests.get(url)
    response.raise_for_status()
    with open(f"{sheet_id}.xlsx", "wb") as f:
        f.write(response.content)


def load_tabs(sheet_id: str) -> op.Workbook:
    return op.load_workbook(f"{sheet_id}.xlsx", keep_links=True)


def clean_name(name: str) -> str:
    if not isinstance(name, str):
        name = str(name)
    for substr in ["\n", "(запись в тг)", "(запись в вк)", "БАН!!"]:
        name = name.replace(substr, "")
    return name.strip()


def get_all_from_tab(ws) -> pd.DataFrame:
    tmp = []
    for i in range(1, ws.max_row):
        row = ws.cell(i, 1)

        if (
            row.value == ""
            or not getattr(row, "hyperlink")
            or row.hyperlink.target == ""
        ):
            continue

        tmp.append({"name": clean_name(row.value), "sources": [row.hyperlink.target]})
    return pd.DataFrame(tmp)


def merge_lists(row):
    def validate_source(src: str) -> list:
        tmp = row.get(src)
        if isinstance(tmp, list):
            return tmp
        return list()

    return sorted(list(set(validate_source("sources_tg") + validate_source("sources_vk"))))


def main(sheet_id: str):
    load_sheet(sheet_id)
    wb = load_tabs(sheet_id)
    tg = get_all_from_tab(wb[TELEGRAM_SHEET])
    vk = get_all_from_tab(wb[VK_SHEET])

    merged = pd.merge(tg, vk, on="name", how="outer", suffixes=("_tg", "_vk"))
    merged["sources"] = merged.apply(merge_lists, axis=1)
    prepared_data = merged[["name", "sources"]]
    
    prepared_data = prepared_data.sort_values("name").reset_index(drop=True)

    with open("output.json", "w", encoding="utf-8") as f:
        json.dump(
            dict(zip(prepared_data["name"], prepared_data["sources"])),
            f,
            ensure_ascii=False,
            indent=2,
        )


if __name__ == "__main__":
    main(os.getenv("GSHEET_ID"))
